import pandas as pd
import matplotlib.pyplot as plt
import os
import subprocess
import platform
from openpyxl import load_workbook
import math
import numpy as np


close_file = "C:/Users/inik2/OneDrive/Рабочий стол/Мои акции/TSLA_close_prices.csv"
data = pd.read_csv(close_file, sep=';', decimal='.')


def calculate_so():
    money = 10000
    period_k = 14 
    period_d = 3
    actions = []
    stocks_amount = 0

    close = data['Close']

    low = close.rolling(window=period_k, min_periods=period_k).min()
    max = close.rolling(window=period_k, min_periods=period_k).max()

    stochastic_k = 100 * ((close - low) / (max - low))

    stochastic_d = stochastic_k.rolling(window=period_d).mean()

    for i in range(period_k+1, len(stochastic_d)):  # Начинаем с i=1, чтобы было предыдущее значение
        # Условия для покупки
        if stochastic_k[i] < 20 and (stochastic_k[i - 1] < stochastic_d[i - 1]) and (stochastic_k[i] > stochastic_d[i]):
            # Проверяем, хватает ли денег на покупку
            if money >= close[i]:
                stocks_to_buy = math.floor(money // close[i])  # Количество акций, которое можем купить
                stocks_amount += stocks_to_buy
                money -= stocks_to_buy * close[i]
                actions.append(f"День {i + 1}: Покупаем {stocks_to_buy} акций по цене {close[i]}, "
                               f"остаток капитала {money:.2f}, "
                               f"%K = {stochastic_k[i]:.2f}, %D = {stochastic_d[i]:.2f}")
            else:
                actions.append(f"День {i + 1}: Недостаточно средств для покупки, остаток капитала {money:.2f}, "
                               f"%K = {stochastic_k[i]:.2f}, %D = {stochastic_d[i]:.2f}")

        # Условия для продажи
        if stochastic_k[i] > 80 and (stochastic_k[i - 1] > stochastic_d[i - 1]) and (stochastic_k[i] < stochastic_d[i]):
            # Продаем все акции, если они есть
            if stocks_amount > 0:
                money += stocks_amount * close[i]
                actions.append(f"День {i + 1}: Продаем все {stocks_amount} акций по цене {close[i]}, "
                               f"остаток капитала {money:.2f}, "
                               f"%K = {stochastic_k[i]:.2f}, %D = {stochastic_d[i]:.2f}")
                stocks_amount = 0
            else:
                actions.append(f"День {i + 1}: Нечего продавать, остаток капитала {money:.2f}, "
                               f"%K = {stochastic_k[i]:.2f}, %D = {stochastic_d[i]:.2f}")
        else:
            actions.append(f"День {i + 1}: Ждём, %K = {stochastic_k[i]:.2f}, %D = {stochastic_d[i]:.2f},"
                           f" капитал {money:.2f}")

    so_DataFrame = pd.DataFrame({

        'Actions': actions
    })


    new_data = pd.DataFrame(so_DataFrame)
    excel_file = 'SO.xlsx'
    sheet_name = 'SO'



    # Создание ExcelWriter с существующим файлом
    try:
        # Проверяем, существует ли файл Excel
        if os.path.exists(excel_file):
            # Если файл существует, загружаем его с помощью load_workbook из openpyxl
            book = load_workbook(excel_file)
        else:
            # Если файл не существует, устанавливаем book в None
            book = None

        # Создаем объект ExcelWriter для записи в файл Excel
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # Если book не равен None (т.е. файл существует), то присоединяем его к writer
            if book:
                writer.book = book
                # Обновляем словарь sheets, чтобы включить существующие листы из загруженной книги
                writer.sheets.update(dict((ws.title, ws) for ws in book.worksheets))

            # Записываем новые данные в указанный лист Excel
            new_data.to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=2, index=False, header=True)

    # Обрабатываем исключения, которые могут возникнуть во время записи
    except Exception as e:
        # Если произошла ошибка, выводим сообщение об ошибке
        print(f"Ошибка при записи в Excel для SO: {e}")

    print(f"Данные успешно добавлены на лист SO в файл {excel_file}.")

    return 




if __name__ == "__main__": 
    calculate_so()