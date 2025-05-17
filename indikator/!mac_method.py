import pandas as pd 
from openpyxl import load_workbook, Workbook, writer   # Библиотека для работы с EXCEL
import os
import matplotlib.pyplot as plt


close_file = "C:/Users/inik2/Desktop/Мои акции/TSLA_close_prices.csv"
data = pd.read_csv(close_file, sep=';', decimal='.')

def mac_calculate():

    period_one = 5  # Период один для расчета
    period_two = 10 # Период два для расчёта 
    close = data['Close']  # Цена закрытия

    one_period_average = close.rolling(window=period_one, min_periods=period_one).mean()
    two_period_average = close.rolling(window=period_two, min_periods=period_two).mean()

    print("Показываю тебе первых 5 значений:", close.head(period_one)) 
    print("Показываю тебе первых 10 значений:", close.head(period_two))



    mac_test_sample = pd.DataFrame({
        'Close': close,
        '5C': one_period_average,
        '10C': two_period_average
    })

    print(mac_test_sample)


    new_data = pd.DataFrame(mac_test_sample)
    excel_file = 'MAC_METHOD.xlsx'
    sheet_name = 'Mac'

    # Создание ExcelWriter с существующим файлом
    try:
        # Проверяем, существует ли файл Excel
        if os.path.exists(excel_file):
            # Если файл существует, загружаем его с помощью load_workbook из openpyxl
            book = load_workbook(excel_file)
        else:
            # Если файл не существует, устанавливаем book в None
            book = None

        # Создаем объект ExcelWriter для записи в файл Excel
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # Если book не равен None (т.е. файл существует), то присоединяем его к writer
            if book:
                writer.book = book
                # Обновляем словарь sheets, чтобы включить существующие листы из загруженной книги
                writer.sheets.update(dict((ws.title, ws) for ws in book.worksheets))

            # Записываем новые данные в указанный лист Excel
            new_data.to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=1, index=False, header=True)

    # Обрабатываем исключения, которые могут возникнуть во время записи
    except Exception as e:
        # Если произошла ошибка, выводим сообщение об ошибке
        print(f"Ошибка при записи в Excel для RW: {e}")

    print(f"Данные успешно добавлены на лист RW в файл {excel_file}.")
    

    plt.figure(figsize=(12, 6)) 
    plt.plot(close, label="Цена закрытия")
    plt.plot(one_period_average, label="5-дневная скользящая средняя", color = "red") 
    plt.plot(two_period_average, label="10-дневная скользящая средняя", color = "green") 
    plt.title(f"Динамика изменения цены на акции {close_file}") 
    plt.xlabel("Дата")
    plt.ylabel("Цена") 
    plt.legend() 
    plt.show()


    return

if __name__ == "__main__":
    mac_calculate()