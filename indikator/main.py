import tkinter as tk #Библиотека для создания интерфейса
from tkinter import filedialog
import pandas as pd #Библиотека для работы с DataFrame
import matplotlib.pyplot as plt #Библиотека для создания графиков
import os #Библиотека для  работы с файловой системой
import subprocess #Библиотека для взаидействия с системой
from openpyxl import load_workbook #Библиотека для работы с EXCEL
import math #Библиотека для написания сложных формул
import numpy as np #Библиотека для работы с большими числами
from  Utilits.Utilits import open_image, open_excel
# from Methods.methods import calculate_rsi, calculate_mac, calculate_so

# Глобальная переменная для хранения пути к файлу
csv_data = {}
file_name_csv = os.path.basename(csv_data)
monthly_data_by_file = {}  # Хранит данные, разделённые по месяцам

# # Функция для сохранения результатов в Excel
# def save_to_excel(results, file_path):
#     with pd.ExcelWriter(file_path, engine="xlsxwriter") as writer:
#         for month, (data, metrics) in results.items():
#             data.to_excel(writer, sheet_name=str(month), index=False)
#             metrics_df = pd.DataFrame([metrics], columns=['Method', 'Final Cash'])
#             metrics_df.to_excel(writer, sheet_name=str(month), startrow=len(data) + 2, index=False)
#     print(f"Результаты сохранены в {file_path}")
#
# # Функция для анализа данных
# def analyze_data(data):
#     monthly_data = split_data_by_month(data)
#     results = {}
#
#     for month, month_data in monthly_data.items():
#         rsi_cash = calculate_rsi(month_data)
#         mac_cash = calculate_mac(month_data)
#         so_cash = calculate_so(month_data)
#
#         results[month] = (month_data, [
#             {'Method': 'RSI', 'Final Cash': rsi_cash},
#             {'Method': 'MAC', 'Final Cash': mac_cash},
#             {'Method': 'SO', 'Final Cash': so_cash},
#         ])
#
#     return results


# Функция для загрузки нескольких CSV файлов
def load_csv():
    global csv_data
    file_paths = filedialog.askopenfilenames(
        title="Выберите CSV файлы",
        filetypes=(("CSV файлы", "*.csv"), ("Все файлы", "*.*"))
    )
    for file_path in file_paths:
        if file_path:
            file_name = os.path.splitext(os.path.basename(file_path))[0]
            data = pd.read_csv(file_path)
            csv_data[file_name] = data
            print(f"Файл {file_name} успешно загружен!")

    # Обновляем статус в GUI
    if csv_data:
        load_csv_label.config(text=f"Загружено файлов: {len(csv_data)}")
    else:
        load_csv_label.config(text="Файлы не загружены")

# Функция для разделения данных всех файлов по месяцам
def split_all_data_by_month():
    global monthly_data_by_file
    if not csv_data:
        print("Сначала загрузите CSV файлы!")
        return

    # Разделение данных по месяцам для всех файлов
    monthly_data_by_file = {file_name: split_data_by_month(data) for file_name, data in csv_data.items()}

    # Обновляем статус в GUI
    print("Все данные успешно разделены по месяцам!")
    result_label.config(text="Данные разделены по месяцам!")
        

def calculate_rsi(data):

    buy_threshold = 30
    sell_threshold = 70
    actions = []
    money = 10000  # Капитал
    stocks_amount = 0  # Количество акций
    period = 10  # Период для расчета
    start_index = 1 # Индекс, с которого начинаем расчет

    close = data['Close'] .round(2) # Цена закрытия

    delta = close.diff() # Разница между закрытиями

    # Рост и падение цены
    gain = delta.where(delta > 0, 0).round(2)
    loss = -delta.where(delta < 0, 0).round(2)

    # Обрезаем до нужного индекса
    gain = gain[start_index:]
    loss = loss[start_index:]

    # Рассчитываем средние только начиная с нужного индекса
    avg_gain = gain.rolling(window=period, min_periods=period).mean().round(2)
    avg_loss = loss.rolling(window=period, min_periods=period).mean().round(2)

    rs = np.where(avg_loss != 0, avg_gain / avg_loss, 0)


    rsi = 100 - (100 / (1 + rs)).round(2)  # Рассчитываем RSI

    RSI_DataFrame_to_Excel_part1 = pd.DataFrame({
        'Close': close,
        'Delta': delta
    })

    RSI_DataFrame_to_Excel_part2 = pd.DataFrame({
        'Gain': gain,
        'Loss': loss,
        'Avg_gain': avg_gain,
        'AVG_loss': avg_loss,
        'RS': rs,
        'RSI': rsi
    })

    for i in range(len(rsi)):
        # Логика покупки и продажи на основе RSI
        if rsi[i] < buy_threshold:
            # Покупаем, если хватает денег
            if money >= close[i+1]:
                stocks_to_buy = math.floor(money // close[i+1])  # Сколько акций можем купить
                stocks_amount += stocks_to_buy  # Увеличиваем количество акций
                money -= stocks_to_buy * close[i+1]  # Списываем деньги за акции
                actions.append(f"День {i + 1}: Покупаем {stocks_to_buy} штук по цене {close[i]}, осталось {money}, "
                               f"RSI: {rsi[i]}")
            else:
                actions.append(f"День {i+1}: Денег недостаточно для покупки, RSI: {rsi[i]}, баланс: {money}")

        elif rsi[i] > sell_threshold and stocks_amount > 0:
            # Продаем все акции, если RSI выше порога продажи
            money += stocks_amount * close[i+1]  # Добавляем деньги от продажи всех акций
            actions.append(f"День {i+1}: Продаем  {stocks_amount} штук по цене {close[i+1]}, RSI: {rsi[i]}, "
                           f"баланс: {money}")
            stocks_amount = 0  # Обнуляем количество акций

        else:
            # Ожидание, если RSI не в зоне покупки или продажи
            actions.append(f"День {i+1}: Ждём, RSI: {rsi[i]}, баланс: {money}")

        # Итоговая сумма после завершения торговли
    final_cash = (money + stocks_amount * close.iloc[-1]).round(2)  # Общая сумма после продажи всех акций


    RSI_DataFrame_to_Excel_part3 = pd.DataFrame({
        'Actions': actions,
        'Final cash': final_cash
    })

    # Построение графиков
    plt.figure(figsize=(12, 6)) 
    plt.plot(gain, color='blue', label='Рост')
    plt.plot(loss, color='green', label='Падение')
    plt.plot(avg_gain, color='red', label='Cредний рост', linestyle = 'dashed')
    plt.plot(avg_loss, color='orange', label='Cреднее падение', linestyle = 'dashed')
    plt.xlabel("Дата")
    plt.ylabel("Средние значения ") 
    plt.legend()  # Легенда графика
    plt.title(f"Динамика изменения цены на акции") 
    plt.savefig("RSI_PLOT.png")  # Сохраняем график в файл
    plt.clf()

    # Запись в EXCEL
    excel_file = (f'RSI_METHOD_{file_name_csv}.xlsx')
    sheet_name = (f'RSI_{file_name_csv}')

    try:
        # Проверяем, существует ли файл Excel
        if os.path.exists(excel_file):
            # Если файл существует, загружаем его с помощью load_workbook из openpyxl
            book = load_workbook(excel_file)
        else:
            # Если файл не существует, устанавливаем book в None
            book = None

        # Создаем объект ExcelWriter для записи в файл Excel
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # Если файл существует, то присоединяем его к writer
            if book:
                writer.book = book
                # Обновляем словарь листов, чтобы включить существующие листы из загруженной книги
                writer.sheets.update(dict((ws.title, ws) for ws in book.worksheets))

            # Записываем первый набор данных в указанное место
            RSI_DataFrame_to_Excel_part1.to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=0, index=False, header=True)

            # Записываем второй набор данных в указанное место
            RSI_DataFrame_to_Excel_part2.to_excel(writer, sheet_name=sheet_name, startrow=1, startcol=3, index=False, header=True)

            # Записываем второй набор данных в указанное место
            RSI_DataFrame_to_Excel_part3.to_excel(writer, sheet_name=sheet_name, startrow=1, startcol=10, index=False, header=True)

    except Exception as e:
        # Если произошла ошибка, выводим сообщение об ошибке
        print(f"Ошибка при записи в Excel: {e}")

    print(f"Данные успешно добавлены на лист '{sheet_name}' в файл '{excel_file}'.")

    RSI_DataFrame_to_result = pd.DataFrame({
        'Name': [file_name_csv],
        'Buy indicator': [buy_threshold],
        'Sell indicator': [sell_threshold],
        'Money': [money],
        # 'Number of shares': [stocks_to_buy],
        'Final cash': [final_cash]
    }).T
    print(RSI_DataFrame_to_result)

    result = (f"{file_name_csv}:{final_cash}")

    return final_cash

def calculate_mac(data):
    actions = []
    money = 10000 # Капитал
    stocks_amount = 0  # Количество акций
    period_one = 5  # Период один для расчета
    period_two = 10 # Период два для расчёта
    close = data['Close'].round(2)  # Цена закрытия

    one_period_average = close.rolling(window=period_one, min_periods=period_one).mean().round(2)
    two_period_average = close.rolling(window=period_two, min_periods=period_two).mean().round(2)

    MAC_DataFrame_to_Excel_part1 = pd.DataFrame({
        'close': close,
        f'{period_one}AV': one_period_average,
        f'{period_two}AV': two_period_average
    })

    # Построение графика
    plt.figure(figsize=(12, 6)) 
    plt.plot(close, color='black', label='Цена закрытия')
    plt.plot(one_period_average, color='red', label=f"{period_one}AV", linestyle = 'dashed')
    plt.plot(two_period_average, color='orange', label=f"{period_two}AV", linestyle = 'dashed')
    plt.xlabel("Дата")
    plt.ylabel("Среднее значение") 
    plt.legend()  # Легенда графика
    plt.title(f"Динамика изменения цены на акции") 
    plt.savefig("MAC_PLOT.png")  # Сохраняем график в файл
    plt.clf()

    for i in range(len(two_period_average)):
        # Пропускаем итерацию, если скользящие средние не определены (NaN)
        if pd.isna(one_period_average[i]) or pd.isna(two_period_average[i]):
            actions.append(f"День {i+1}: Недостаточно данных для расчета средних")
            continue

        # Логика покупки и продажи
        if two_period_average[i] < one_period_average[i]:
            if money >= close[i]:
                stocks_to_buy = math.floor(money // close[i])  # Сколько акций можно купить
                stocks_amount += stocks_to_buy
                money -= stocks_to_buy * close[i]
                actions.append(f"День {i+1}: Покупаем {stocks_to_buy} штук по цене {close[i]}, осталось {money}, "
                               f"{period_one}AV: {one_period_average[i]}, {period_two}AV:{two_period_average[i]}")
            else:
                actions.append(f"День {i+1}: Недостаточно средств для покупки, баланс: {money}, "
                               f"{period_one}AV: {one_period_average[i]}, {period_two}AV: {two_period_average[i]}")

        if two_period_average[i] > one_period_average[i]:
            if stocks_amount > 0:
                money += stocks_amount * close[i]
                actions.append(f"День {i+1}: Продаём {stocks_amount} штук по цене {close[i]}, баланс: {money}, "
                               f"{period_one}AV: {one_period_average[i]}, {period_two}AV: {two_period_average[i]}")
                stocks_amount = 0
            else:
                actions.append(f"День {i + 1}: Нечего продавать, баланс: {money}, "
                               f"{period_one}AV: {one_period_average[i]}, {period_two}AV: {two_period_average[i]}")

    # Финальный расчет капитала
    final_cash = (money + stocks_amount * close.iloc[-1]).round(2) # Общая сумма по текущей стоимости акций

    MAC_DataFrame_to_Excel_part2 = pd.DataFrame({'actions': actions})

    excel_file = 'MAC_METHOD.xlsx'
    sheet_name = 'MAC'

    # Создание ExcelWriter с существующим файлом
    try:
        # Проверяем, существует ли файл Excel
        if os.path.exists(excel_file):
            # Если файл существует, загружаем его с помощью load_workbook из openpyxl
            book = load_workbook(excel_file)
        else:
            # Если файл не существует, устанавливаем book в None
            book = None

        # Создаем объект ExcelWriter для записи в файл Excel
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # Если book не равен None (т.е. файл существует), то присоединяем его к writer
            if book:
                writer.book = book
                # Обновляем словарь sheets, чтобы включить существующие листы из загруженной книги
                writer.sheets.update(dict((ws.title, ws) for ws in book.worksheets))

            # Записываем новые данные в указанный лист Excel
            MAC_DataFrame_to_Excel_part1.to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=0, index=False,
                                                  header=True)

            MAC_DataFrame_to_Excel_part2.to_excel(writer, sheet_name=sheet_name, startrow=1, startcol=4, index=False,
                                                  header=False)

    # Обрабатываем исключения, которые могут возникнуть во время записи
    except Exception as e:
        # Если произошла ошибка, выводим сообщение об ошибке
        print(f"Ошибка при записи в Excel: {e}")

    print(f"Данные успешно добавлены на лист MAC в файл {excel_file}.")

    MAC_DataFrame_to_result = pd.DataFrame({
        'Name': [file_name_csv],
        'Small period': [period_one],
        'Long period': [period_two],
        'Money': [money],
        # 'Number of shares': [stocks_to_buy],
        'Final cash': [final_cash]
    }).T
    print(MAC_DataFrame_to_result)

    result = (f"{file_name_csv}:{final_cash}")

    return final_cash

def calculate_so(data):
    money = 10000
    low_parameter = 20
    high_parameter = 80
    period_k = 14
    period_d = 3
    actions = []
    stocks_amount = 0

    close = data['Close']

    low = close.rolling(window=period_k, min_periods=period_k).min()
    max = close.rolling(window=period_k, min_periods=period_k).max()

    stochastic_k = 100 * ((close - low) / (max - low))

    stochastic_d = stochastic_k.rolling(window=period_d).mean()


    SO_DataFrame_to_Excel_part1 = pd.DataFrame({
        'Close': close,
        'Low': low,
        'Max': max,
        '%K': stochastic_k,
        '%D': stochastic_d
    })

    for i in range(period_k + 1, len(stochastic_d)):  # Начинаем с i=1, чтобы было предыдущее значение
        # Условия для покупки
        if stochastic_k[i] < low_parameter and (stochastic_k[i - 1] < stochastic_d[i - 1]) \
                and (stochastic_k[i] > stochastic_d[i]):
            # Проверяем, хватает ли денег на покупку
            if money >= close[i]:
                stocks_to_buy = math.floor(money // close[i])  # Количество акций, которое можем купить
                stocks_amount += stocks_to_buy
                money -= stocks_to_buy * close[i]
                actions.append(f"День {i + 1}: Покупаем {stocks_to_buy} штук по цене {close[i]}, осталось {money}, "
                               f"%K: {stochastic_k[i]}, %D: {stochastic_d[i]}, "
                               f"%K-1: {stochastic_k[i-1]}, %D-1: {stochastic_d[i-1]} ")
            else:
                actions.append(f"День {i + 1}: Недостаточно средств для покупки, баланс: {money}, "
                               f"%K: {stochastic_k[i]}, %D: {stochastic_d[i]}, "
                               f"%K-1: {stochastic_k[i-1]}, %D-1: {stochastic_d[i-1]}")

        # Условия для продажи
        if stochastic_k[i] > high_parameter and (stochastic_k[i - 1] > stochastic_d[i - 1]) \
                and (stochastic_k[i] < stochastic_d[i]):
            # Продаем все акции, если они есть
            if stocks_amount > 0:
                money += stocks_amount * close[i]
                actions.append(f"День {i + 1}: Продаем {stocks_amount} штук по цене {close[i]}, баланс: {money}, "
                               f"%K: {stochastic_k[i]}, %D: {stochastic_d[i]}" 
                               f"%K-1: {stochastic_k[i-1]}, %D-1: {stochastic_d[i-1]}")
                stocks_amount = 0
            else:
                actions.append(f"День {i + 1}: Нечего продавать, баланс: {money}, "
                               f"%K: {stochastic_k[i]}, %D: {stochastic_d[i]}"
                               f"%K-1: {stochastic_k[i - 1]}, %D-1: {stochastic_d[i - 1]}")
        else:
            actions.append(f"День {i + 1}: Ждём, %K: {stochastic_k[i]}, %D: {stochastic_d[i]},"
                           f"%K-1: {stochastic_k[i - 1]}, %D-1: {stochastic_d[i - 1]}"
                           f"баланс: {money}")


    final_cash = (money + stocks_amount * close.iloc[-1]).round(2)

    SO_DataFrame_to_Excel_part2 = pd.DataFrame({'actions': actions})


    plt.figure(figsize=(12, 6)) 
    plt.plot(close, color='black', label='Цена закрытия')
    plt.plot(low, color='red', label='Минимальное значение', linestyle = 'dashed')
    plt.plot(max, color='orange', label='Максимальное значение', linestyle = 'dashed')
    plt.plot(stochastic_k, color='blue', label='К', linestyle = 'solid')
    plt.plot(stochastic_d, color='green', label='Д', linestyle = 'solid')
    plt.xlabel("Дата")
    plt.ylabel("Среднее значение") 
    plt.legend()  # Легенда графика
    plt.title(f"{file_name_csv}")
    plt.savefig("SO_PLOT.png")  # Сохраняем график в файл
    plt.clf() 


    excel_file = 'SO_METHOD.xlsx'
    sheet_name = 'SO'

    # Создание ExcelWriter с существующим файлом
    try:
        # Проверяем, существует ли файл Excel
        if os.path.exists(excel_file):
            # Если файл существует, загружаем его с помощью load_workbook из openpyxl
            book = load_workbook(excel_file)
        else:
            # Если файл не существует, устанавливаем book в None
            book = None

        # Создаем объект ExcelWriter для записи в файл Excel
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # Если book не равен None (т.е. файл существует), то присоединяем его к writer
            if book:
                writer.book = book
                # Обновляем словарь sheets, чтобы включить существующие листы из загруженной книги
                writer.sheets.update(dict((ws.title, ws) for ws in book.worksheets))

            # Записываем новые данные в указанный лист Excel
            SO_DataFrame_to_Excel_part1.to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=0, index=False,
                                                  header=True)
            SO_DataFrame_to_Excel_part2.to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=6, index=False,
                                                  header=True)

    # Обрабатываем исключения, которые могут возникнуть во время записи
    except Exception as e:
        # Если произошла ошибка, выводим сообщение об ошибке
        print(f"Ошибка при записи в Excel: {e}")

    print(f"Данные успешно добавлены на лист SO в файл {excel_file}.")

    SO_DataFrame_to_result = pd.DataFrame({
        'Name': [file_name_csv],
        'Period k': [period_k],
        'Period d': [period_d],
        'Low parameter': [low_parameter],
        'high_parameter': [high_parameter],
        'Money': [money],
        # 'Number of shares': [stocks_to_buy],
        'Final cash': [final_cash]
    }).T
    print(SO_DataFrame_to_result)

    result = (f"{file_name_csv}:{final_cash}")

    return final_cash

# Функция для обновления интерфейса
def run_analysis():
    global csv_file_path

    if not csv_file_path:
        result_label.config(text="Сначала загрузите CSV файл.")
        return

    try:
        # Чтение данных
        data = pd.read_csv(csv_file_path)

        # Выполняем анализ
        analysis_results = analyze_data(data)

        # Сохраняем результаты
        save_to_excel(analysis_results, "output_analysis.xlsx")

        # Обновляем интерфейс
        result_label.config(text="Анализ завершен. Результаты сохранены в output_analysis.xlsx")

    except Exception as e:
        result_label.config(text=f"Ошибка анализа: {str(e)}")

# Класс для создания всплывающих подсказок
class Tooltip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tooltip_window = None
        self.widget.bind("<Enter>", self.show_tooltip)
        self.widget.bind("<Leave>", self.hide_tooltip)

    def show_tooltip(self, event=None):
        if self.tooltip_window is not None:
            return

        # Создаем окно для подсказки
        self.tooltip_window = tk.Toplevel(self.widget)
        self.tooltip_window.wm_overrideredirect(True)
        self.tooltip_window.wm_geometry(f"+{event.x_root + 10}+{event.y_root + 10}")

        # Добавляем текст подсказки в окно
        label = tk.Label(self.tooltip_window, text=self.text, background="black", fg="white", relief="solid",
                         borderwidth=1, font=("Courier", 11, "bold"))
        label.pack()

    def hide_tooltip(self, event=None):
        if self.tooltip_window:
            self.tooltip_window.destroy()
            self.tooltip_window = None

# Создание интерфейса с терминальным стилем
def create_interface():
    global load_csv_label, rsi_label, mac_label, so_label, result_label
    global rsi_image_btn, rsi_excel_btn, mac_image_btn, mac_excel_btn, so_image_btn, so_excel_btn

    # Создание основного окна
    root = tk.Tk()
    root.title("Анализ CSV файла")
    root.geometry("1200x680")
    root.configure(bg="black")  # Черный фон

    # Настройки стиля для терминала
    terminal_font = ("Courier", 11, "bold")  # Шрифт фиксированной ширины
    fg_color = "White"  # Белый текст

    # Верхний фрейм для загрузки файла
    frame_top = tk.Frame(root, bd=2, relief="raised", bg="grey")
    frame_top.pack(fill="x", pady=10)

    # Кнопка для загрузки CSV файла
    load_csv_btn = tk.Button(frame_top, text="Загрузка CSV файла", width=20, command=load_csv,
                             font=terminal_font, fg=fg_color, bg="black", relief="groove")

    load_csv_btn.pack(side="left", padx=100, pady=10)

    # Метка для отображения статуса загрузки
    load_csv_label = tk.Label(frame_top, text="Файл не загружен", anchor="w", font=terminal_font,
                              fg=fg_color, bg="black", relief="groove")

    load_csv_label.pack(side="right", padx=50)

    # ФРЕЙМ РЕЗУЛЬТАТОВ
    frame_middle_part_1 = tk.Frame(root, bd=2, relief="solid", bg="black")
    frame_middle_part_1.pack(fill="x", pady=10)

    rsi_label = tk.Label(frame_middle_part_1, text="RSI", width=30, height=15, font=terminal_font, fg=fg_color,
                         bg="black", bd=2, relief="groove")
    rsi_label.pack(side="right")
    Tooltip(rsi_label, "Индикатор RSI: Показатель силы тренда")

    mac_label = tk.Label(frame_middle_part_1, text="MAC", width=30, height=15, font=terminal_font, fg=fg_color,
                         bg="black", bd=2, relief="groove")
    mac_label.pack(side="left")
    Tooltip(mac_label, "Индикатор MAC: Показатель скользящего среднего")

    so_label = tk.Label(frame_middle_part_1, text="SO", width=30, height=15, font=terminal_font,
                        fg=fg_color, bg="black", bd=2, relief="groove")
    so_label.pack(side="right", padx=185)
    Tooltip(so_label, "Индикатор SO: Стохастический осциллятор")

    # ФРЕЙМ КНОКИ ГРАФИКА
    frame_middle_part_2 = tk.Frame(root, bd=2, relief="solid", bg="black")
    frame_middle_part_2.pack(fill="x", pady=10)

    rsi_image_btn = tk.Button(frame_middle_part_2, text="График", width=30,height=2, font=terminal_font, fg=fg_color,
                         bg="black", bd=2, relief="groove", state="disabled")
    rsi_image_btn.pack(side="right")

    mac_image_btn = tk.Button(frame_middle_part_2, text="График",width=30, height=2, font=terminal_font, fg=fg_color,
                              bg="black", bd=2, relief="groove", state="disabled")
    mac_image_btn.pack(side="left")

    so_image_btn = tk.Button(frame_middle_part_2, text="График",width=30, height=2, font=terminal_font, fg=fg_color,
                              bg="black", bd=2, relief="groove", state="disabled")
    so_image_btn.pack(side="right", padx=175)

    # ФРЕЙМ КНОКИ EXCEL
    frame_middle_part_3 = tk.Frame(root, bd=2, relief="solid", bg="black")
    frame_middle_part_3.pack(fill="x", pady=10)

    rsi_excel_btn = tk.Button(frame_middle_part_3, text="EXCEL", width=30, height=2, font=terminal_font, fg=fg_color,
                              bg="black", bd=2, relief="groove", state="disabled")
    rsi_excel_btn.pack(side="right")

    mac_excel_btn = tk.Button(frame_middle_part_3, text="EXCEL", width=30, height=2, font=terminal_font, fg=fg_color,
                              bg="black", bd=2, relief="groove", state="disabled")
    mac_excel_btn.pack(side="left")

    so_excel_btn = tk.Button(frame_middle_part_3, text="EXCEL", width=30, height=2, font=terminal_font, fg=fg_color,
                             bg="black", bd=2, relief="groove", state="disabled")
    so_excel_btn.pack(side="right", padx=175)

    # Нижний фрейм с кнопкой анализа и результатом
    frame_bottom_part_1 = tk.Frame(root, bd=2, relief="flat", bg="black")
    frame_bottom_part_1.pack(fill="x", pady=10)

    analyze_btn = tk.Button(frame_bottom_part_1, text="Анализ", width=30, height=2, command=run_analysis, font=terminal_font,
                            fg=fg_color, bg="black", relief="groove")
    analyze_btn.pack(pady=25)

    # Метка для вывода возможных ошибок или статуса
    result_label = tk.Label(frame_bottom_part_1, text="", anchor="w", font=terminal_font, fg=fg_color, bg="black")
    result_label.pack(pady=10)



    # Запуск главного цикла приложения
    root.resizable(False, False)
    root.mainloop()

if __name__ == "__main__":
    create_interface()

